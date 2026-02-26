"""
Cross-Tenant Partnerships & Transfers Router.
Endpoint: /api/v1/{tenant_slug}/partners/...
"""

from typing import Optional
from fastapi import APIRouter, Depends, HTTPException, status
from sqlalchemy.orm import Session
from pydantic import BaseModel

from database import get_db
from database.models import User
from core.dependencies import get_current_active_user
from core.tenant import get_current_tenant_id
from services.cross_transfer import PartnershipService, CrossTransferService


router = APIRouter()


# ==================== PARTNERSHIPS ====================

@router.get("/search")
async def search_companies(
    q: str,
    current_user: User = Depends(get_current_active_user),
    db: Session = Depends(get_db)
):
    """Search for companies to partner with."""
    tenant_id = get_current_tenant_id()
    service = PartnershipService(db)
    results = service.search_tenants(q, tenant_id)
    return {"data": results}


@router.get("")
async def list_partners(
    current_user: User = Depends(get_current_active_user),
    db: Session = Depends(get_db)
):
    """List all my partnerships."""
    tenant_id = get_current_tenant_id()
    service = PartnershipService(db)
    partners = service.get_my_partners(tenant_id)
    return {"data": partners}


class PartnerRequestBody(BaseModel):
    target_tenant_id: int
    notes: Optional[str] = None


@router.post("/request")
async def send_partner_request(
    body: PartnerRequestBody,
    current_user: User = Depends(get_current_active_user),
    db: Session = Depends(get_db)
):
    """Send partnership request to another company."""
    tenant_id = get_current_tenant_id()
    service = PartnershipService(db)
    ok, msg = service.send_request(tenant_id, body.target_tenant_id, body.notes)
    if not ok:
        raise HTTPException(400, msg)
    return {"success": True, "message": msg}


@router.post("/{partnership_id}/accept")
async def accept_partnership(
    partnership_id: int,
    current_user: User = Depends(get_current_active_user),
    db: Session = Depends(get_db)
):
    """Accept partnership request."""
    tenant_id = get_current_tenant_id()
    service = PartnershipService(db)
    ok, msg = service.accept_request(partnership_id, tenant_id)
    if not ok:
        raise HTTPException(400, msg)
    return {"success": True, "message": msg}


@router.post("/{partnership_id}/reject")
async def reject_partnership(
    partnership_id: int,
    current_user: User = Depends(get_current_active_user),
    db: Session = Depends(get_db)
):
    """Reject partnership request."""
    tenant_id = get_current_tenant_id()
    service = PartnershipService(db)
    ok, msg = service.reject_request(partnership_id, tenant_id)
    if not ok:
        raise HTTPException(400, msg)
    return {"success": True, "message": msg}


@router.delete("/{partnership_id}")
async def remove_partnership(
    partnership_id: int,
    current_user: User = Depends(get_current_active_user),
    db: Session = Depends(get_db)
):
    """Remove partnership."""
    tenant_id = get_current_tenant_id()
    service = PartnershipService(db)
    ok, msg = service.remove_partner(partnership_id, tenant_id)
    if not ok:
        raise HTTPException(400, msg)
    return {"success": True, "message": msg}


# ==================== CROSS-TENANT TRANSFERS ====================

class TransferItemBody(BaseModel):
    product_id: int
    quantity: float
    uom_id: int
    sale_price: float
    sale_price_usd: Optional[float] = None
    notes: Optional[str] = None


class CreateTransferBody(BaseModel):
    receiver_tenant_id: int
    sender_warehouse_id: int
    items: list[TransferItemBody]
    notes: Optional[str] = None


@router.post("/transfers")
async def create_cross_transfer(
    body: CreateTransferBody,
    current_user: User = Depends(get_current_active_user),
    db: Session = Depends(get_db)
):
    """Send products to partner company."""
    tenant_id = get_current_tenant_id()
    service = CrossTransferService(db)

    items = [i.model_dump() for i in body.items]

    try:
        transfer, msg = service.create_transfer(
            sender_tenant_id=tenant_id,
            sender_warehouse_id=body.sender_warehouse_id,
            sender_user_id=current_user.id,
            receiver_tenant_id=body.receiver_tenant_id,
            items=items,
            notes=body.notes,
        )
        if not transfer:
            raise HTTPException(400, msg)
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(500, f"Transfer yaratishda xatolik: {str(e)}")

    return {
        "success": True,
        "message": msg,
        "data": {
            "id": transfer.id,
            "transfer_number": transfer.transfer_number,
        }
    }


@router.get("/transfers/outgoing")
async def get_outgoing_transfers(
    current_user: User = Depends(get_current_active_user),
    db: Session = Depends(get_db)
):
    """Get transfers I sent."""
    tenant_id = get_current_tenant_id()
    service = CrossTransferService(db)
    return {"data": service.get_outgoing(tenant_id)}


@router.get("/transfers/incoming")
async def get_incoming_transfers(
    current_user: User = Depends(get_current_active_user),
    db: Session = Depends(get_db)
):
    """Get transfers sent to me."""
    tenant_id = get_current_tenant_id()
    service = CrossTransferService(db)
    return {"data": service.get_incoming(tenant_id)}


@router.get("/transfers/{transfer_id}")
async def get_transfer_detail(
    transfer_id: int,
    current_user: User = Depends(get_current_active_user),
    db: Session = Depends(get_db)
):
    """Get transfer details."""
    tenant_id = get_current_tenant_id()
    service = CrossTransferService(db)
    detail = service.get_transfer_detail(transfer_id, tenant_id)
    if not detail:
        raise HTTPException(404, "Transfer topilmadi")
    return {"data": detail}


class ProductMapping(BaseModel):
    """Map incoming product to existing product in receiver's catalog."""
    transfer_item_id: int
    receiver_product_id: Optional[int] = None  # None = auto-create new product


class AcceptTransferBody(BaseModel):
    warehouse_id: int
    mappings: Optional[list[ProductMapping]] = None  # If not provided, auto-create all


@router.post("/transfers/{transfer_id}/accept")
async def accept_cross_transfer(
    transfer_id: int,
    body: AcceptTransferBody,
    current_user: User = Depends(get_current_active_user),
    db: Session = Depends(get_db)
):
    """Accept incoming transfer and add to warehouse."""
    tenant_id = get_current_tenant_id()
    service = CrossTransferService(db)
    ok, msg = service.accept_transfer(
        transfer_id=transfer_id,
        receiver_tenant_id=tenant_id,
        receiver_user_id=current_user.id,
        receiver_warehouse_id=body.warehouse_id,
        product_mappings={m.transfer_item_id: m.receiver_product_id for m in body.mappings} if body.mappings else None,
    )
    if not ok:
        raise HTTPException(400, msg)
    return {"success": True, "message": msg}


class EditTransferItemBody(BaseModel):
    item_id: int
    sale_price: float


class EditTransferBody(BaseModel):
    items: list[EditTransferItemBody]
    notes: Optional[str] = None


@router.put("/transfers/{transfer_id}/edit")
async def edit_cross_transfer(
    transfer_id: int,
    body: EditTransferBody,
    current_user: User = Depends(get_current_active_user),
    db: Session = Depends(get_db)
):
    """Edit transfer items (prices). Both parties can edit."""
    tenant_id = get_current_tenant_id()
    service = CrossTransferService(db)
    items_updates = [{"item_id": i.item_id, "sale_price": i.sale_price} for i in body.items]
    ok, msg = service.edit_transfer(transfer_id, tenant_id, items_updates, body.notes)
    if not ok:
        raise HTTPException(400, msg)
    return {"success": True, "message": msg}


@router.post("/transfers/{transfer_id}/confirm")
async def confirm_transfer_edit(
    transfer_id: int,
    current_user: User = Depends(get_current_active_user),
    db: Session = Depends(get_db)
):
    """Confirm other party's edit. Sets last_edited_by to you."""
    tenant_id = get_current_tenant_id()
    service = CrossTransferService(db)
    ok, msg = service.confirm_edit(transfer_id, tenant_id)
    if not ok:
        raise HTTPException(400, msg)
    return {"success": True, "message": msg}


class RejectTransferBody(BaseModel):
    reason: Optional[str] = None


@router.post("/transfers/{transfer_id}/reject")
async def reject_cross_transfer(
    transfer_id: int,
    body: RejectTransferBody = RejectTransferBody(),
    current_user: User = Depends(get_current_active_user),
    db: Session = Depends(get_db)
):
    """Reject incoming transfer (stock returned to sender)."""
    tenant_id = get_current_tenant_id()
    service = CrossTransferService(db)
    ok, msg = service.reject_transfer(transfer_id, tenant_id, body.reason)
    if not ok:
        raise HTTPException(400, msg)
    return {"success": True, "message": msg}


# ==================== EXCEL EXPORT ====================

@router.get("/transfers/{transfer_id}/export")
async def export_transfer_excel(
    transfer_id: int,
    current_user: User = Depends(get_current_active_user),
    db: Session = Depends(get_db)
):
    """Export transfer details as Excel file."""
    from fastapi.responses import StreamingResponse
    import openpyxl
    from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
    from io import BytesIO

    tenant_id = get_current_tenant_id()
    service = CrossTransferService(db)
    detail = service.get_transfer_detail(transfer_id, tenant_id)
    if not detail:
        raise HTTPException(404, "Transfer topilmadi")

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Transfer"

    # Styles
    header_font = Font(bold=True, size=14)
    sub_font = Font(bold=True, size=11)
    bold = Font(bold=True)
    thin_border = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin')
    )
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font_white = Font(bold=True, color="FFFFFF", size=10)

    # Title
    ws.merge_cells('A1:F1')
    ws['A1'] = f"Transfer hisoboti: {detail['transfer_number']}"
    ws['A1'].font = header_font

    # Info
    info = [
        ("Sana:", detail['transfer_date']),
        ("Holat:", detail['status'].upper()),
        ("Yuboruvchi:", f"{detail['sender_tenant_name']} ({detail['sender_warehouse_name']})"),
        ("Qabul qiluvchi:", f"{detail['receiver_tenant_name']}" + (f" ({detail['receiver_warehouse_name']})" if detail.get('receiver_warehouse_name') else "")),
    ]
    if detail.get('notes'):
        info.append(("Izoh:", detail['notes']))

    for idx, (label, value) in enumerate(info, start=3):
        ws[f'A{idx}'] = label
        ws[f'A{idx}'].font = bold
        ws[f'B{idx}'] = value

    # Table header
    table_start = len(info) + 5
    headers = ["#", "Mahsulot", "Miqdor", "O'lchov", "Narx (so'm)", "Jami (so'm)"]
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=table_start, column=col, value=h)
        cell.font = header_font_white
        cell.fill = header_fill
        cell.border = thin_border
        cell.alignment = Alignment(horizontal='center')

    # Items
    for i, item in enumerate(detail['items'], 1):
        row = table_start + i
        vals = [i, item['product_name'], item['quantity'], item['uom_symbol'], item['sale_price'], item['total_amount']]
        for col, v in enumerate(vals, 1):
            cell = ws.cell(row=row, column=col, value=v)
            cell.border = thin_border
            if col >= 3:
                cell.number_format = '#,##0.##'
                cell.alignment = Alignment(horizontal='right')

    # Total row
    total_row = table_start + len(detail['items']) + 1
    ws.merge_cells(f'A{total_row}:D{total_row}')
    ws[f'A{total_row}'] = "JAMI:"
    ws[f'A{total_row}'].font = bold
    ws[f'A{total_row}'].alignment = Alignment(horizontal='right')
    ws[f'F{total_row}'] = detail['total_amount']
    ws[f'F{total_row}'].font = bold
    ws[f'F{total_row}'].number_format = '#,##0.##'
    ws[f'F{total_row}'].border = thin_border

    # Column widths
    ws.column_dimensions['A'].width = 5
    ws.column_dimensions['B'].width = 35
    ws.column_dimensions['C'].width = 12
    ws.column_dimensions['D'].width = 10
    ws.column_dimensions['E'].width = 15
    ws.column_dimensions['F'].width = 18

    # Save
    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)

    filename = f"transfer_{detail['transfer_number']}.xlsx"
    return StreamingResponse(
        buffer,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'}
    )


# ==================== STATISTICS ====================

@router.get("/stats/summary")
async def get_partners_summary(
    current_user: User = Depends(get_current_active_user),
    db: Session = Depends(get_db)
):
    """Get summary stats for all partners."""
    from services.cross_transfer import PartnerStatsService
    tenant_id = get_current_tenant_id()
    service = PartnerStatsService(db)
    try:
        return {"data": service.get_all_partners_summary(tenant_id)}
    except Exception as e:
        import traceback
        traceback.print_exc()
        raise HTTPException(500, f"Stats xatolik: {str(e)}")


@router.get("/stats/{partner_id}")
async def get_partner_stats(
    partner_id: int,
    current_user: User = Depends(get_current_active_user),
    db: Session = Depends(get_db)
):
    """Get detailed stats for a specific partner."""
    from services.cross_transfer import PartnerStatsService
    tenant_id = get_current_tenant_id()
    service = PartnerStatsService(db)
    try:
        return {"data": service.get_partner_stats(tenant_id, partner_id)}
    except Exception as e:
        import traceback
        traceback.print_exc()
        raise HTTPException(500, f"Stats xatolik: {str(e)}")


# ==================== PAYMENTS ====================

class AddPaymentBody(BaseModel):
    partner_tenant_id: int
    amount: float
    payment_type: str = "cash"
    notes: Optional[str] = None
    direction: str = "outgoing"  # outgoing = I pay them, incoming = they pay me


@router.post("/payments")
async def add_payment(
    body: AddPaymentBody,
    current_user: User = Depends(get_current_active_user),
    db: Session = Depends(get_db)
):
    """Record a payment between partners."""
    from services.cross_transfer import PartnerPaymentService
    tenant_id = get_current_tenant_id()
    service = PartnerPaymentService(db)

    if body.direction == "outgoing":
        payer = tenant_id
        receiver = body.partner_tenant_id
    else:
        payer = body.partner_tenant_id
        receiver = tenant_id

    ok, msg = service.add_payment(
        payer_tenant_id=payer,
        receiver_tenant_id=receiver,
        amount=body.amount,
        payment_type=body.payment_type,
        notes=body.notes,
        created_by_id=current_user.id,
    )
    if not ok:
        raise HTTPException(400, msg)
    return {"success": True, "message": msg}


@router.get("/payments/{partner_id}")
async def get_payments(
    partner_id: int,
    current_user: User = Depends(get_current_active_user),
    db: Session = Depends(get_db)
):
    """Get payment history with a partner."""
    from services.cross_transfer import PartnerPaymentService
    tenant_id = get_current_tenant_id()
    service = PartnerPaymentService(db)
    return {"data": service.get_payments(tenant_id, partner_id)}


@router.post("/payments/{payment_id}/confirm")
async def confirm_payment(
    payment_id: int,
    current_user: User = Depends(get_current_active_user),
    db: Session = Depends(get_db)
):
    """Confirm a pending payment."""
    from services.cross_transfer import PartnerPaymentService
    tenant_id = get_current_tenant_id()
    service = PartnerPaymentService(db)
    ok, msg = service.confirm_payment(payment_id, tenant_id, current_user.id)
    if not ok:
        raise HTTPException(400, msg)
    return {"success": True, "message": msg}


@router.post("/payments/{payment_id}/reject")
async def reject_payment(
    payment_id: int,
    body: RejectTransferBody = RejectTransferBody(),
    current_user: User = Depends(get_current_active_user),
    db: Session = Depends(get_db)
):
    """Reject a pending payment."""
    from services.cross_transfer import PartnerPaymentService
    tenant_id = get_current_tenant_id()
    service = PartnerPaymentService(db)
    ok, msg = service.reject_payment(payment_id, tenant_id, body.reason)
    if not ok:
        raise HTTPException(400, msg)
    return {"success": True, "message": msg}


# ==================== NOTIFICATIONS ====================

@router.get("/notifications")
async def get_notifications(
    current_user: User = Depends(get_current_active_user),
    db: Session = Depends(get_db)
):
    """Get notifications for current tenant."""
    from services.cross_transfer import NotificationService
    tenant_id = get_current_tenant_id()
    service = NotificationService(db)
    return {
        "data": service.get_notifications(tenant_id),
        "unread_count": service.get_unread_count(tenant_id),
    }


@router.get("/notifications/count")
async def get_notification_count(
    current_user: User = Depends(get_current_active_user),
    db: Session = Depends(get_db)
):
    """Get unread notification count (lightweight for header polling)."""
    from services.cross_transfer import NotificationService
    tenant_id = get_current_tenant_id()
    service = NotificationService(db)
    return {"unread_count": service.get_unread_count(tenant_id)}


@router.post("/notifications/read")
async def mark_notifications_read(
    current_user: User = Depends(get_current_active_user),
    db: Session = Depends(get_db)
):
    """Mark all notifications as read."""
    from services.cross_transfer import NotificationService
    tenant_id = get_current_tenant_id()
    service = NotificationService(db)
    service.mark_read(tenant_id)
    return {"success": True}


@router.post("/notifications/{notification_id}/read")
async def mark_notification_read(
    notification_id: int,
    current_user: User = Depends(get_current_active_user),
    db: Session = Depends(get_db)
):
    """Mark single notification as read."""
    from services.cross_transfer import NotificationService
    tenant_id = get_current_tenant_id()
    service = NotificationService(db)
    service.mark_read(tenant_id, notification_id)
    return {"success": True}
