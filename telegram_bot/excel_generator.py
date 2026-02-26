"""
Excel Generator for Telegram Bot — Multi-Tenant SaaS
Generates professional Excel files for purchase, payment, and daily report notifications.
Tenant name is passed per-request — no hardcoded values.
"""
import io
from datetime import datetime, date
from typing import List, Dict, Any
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill


class ExcelGenerator:
    """Generates Excel files with tenant branding."""

    HEADER_FONT = Font(bold=True, color="FFFFFF", size=11)
    HEADER_FILL = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    TITLE_FONT = Font(bold=True, size=12)
    BOLD_FONT = Font(bold=True, size=11)
    BORDER = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin')
    )
    SUCCESS_FILL = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
    WARNING_FILL = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
    CENTER = Alignment(horizontal='center', vertical='center')
    RIGHT = Alignment(horizontal='right', vertical='center')

    def _fmt(self, val: Any) -> str:
        try:
            return f"{float(val or 0):,.0f}".replace(",", " ")
        except:
            return str(val)

    def _save(self, wb: Workbook) -> bytes:
        buf = io.BytesIO()
        wb.save(buf)
        buf.seek(0)
        return buf.getvalue()

    def _header_row(self, ws, row: int, headers: List[str]):
        for c, h in enumerate(headers, 1):
            cell = ws.cell(row=row, column=c, value=h)
            cell.font = self.HEADER_FONT
            cell.fill = self.HEADER_FILL
            cell.border = self.BORDER
            cell.alignment = self.CENTER

    # ==================== PURCHASE EXCEL ====================

    def generate_purchase(
        self, tenant_name: str, customer_name: str, customer_phone: str,
        sale_number: str, sale_date, items: List[Dict], total: float,
        paid: float, debt: float, operator: str
    ) -> bytes:
        try:
            wb = Workbook()
            ws = wb.active
            ws.title = "Harid"

            ws.merge_cells('A1:F1')
            ws['A1'] = f"📦 HARID CHEKI — {tenant_name}"
            ws['A1'].font = Font(bold=True, size=14)
            ws['A1'].alignment = self.CENTER

            date_str = sale_date.strftime('%d.%m.%Y %H:%M') if isinstance(sale_date, datetime) else str(sale_date)
            info = [("Mijoz:", customer_name), ("Telefon:", customer_phone),
                    ("Chek:", sale_number), ("Sana:", date_str), ("Kassir:", operator)]
            for i, (k, v) in enumerate(info, 3):
                ws[f'A{i}'] = k; ws[f'A{i}'].font = self.TITLE_FONT
                ws[f'B{i}'] = v

            row = 9
            self._header_row(ws, row, ['№', 'Tovar', 'Miqdor', 'Narx', 'Chegirma', 'Jami'])
            for idx, item in enumerate(items, 1):
                row += 1
                ws.cell(row=row, column=1, value=idx).border = self.BORDER
                ws.cell(row=row, column=2, value=item.get('product_name', '')).border = self.BORDER
                ws.cell(row=row, column=3, value=f"{item.get('quantity', 0)} {item.get('uom_symbol', '')}").border = self.BORDER
                ws.cell(row=row, column=4, value=self._fmt(item.get('unit_price', 0))).border = self.BORDER
                ws.cell(row=row, column=5, value=self._fmt(item.get('discount_amount', 0))).border = self.BORDER
                c = ws.cell(row=row, column=6, value=self._fmt(item.get('total_price', 0)))
                c.border = self.BORDER; c.font = self.BOLD_FONT

            row += 2
            for label, amt, fill in [("JAMI:", total, None), ("TO'LANGAN:", paid, self.SUCCESS_FILL),
                                      ("QARZ:", debt, self.WARNING_FILL if debt > 0 else self.SUCCESS_FILL)]:
                ws.merge_cells(f'A{row}:E{row}')
                ws[f'A{row}'] = label; ws[f'A{row}'].font = self.TITLE_FONT; ws[f'A{row}'].alignment = self.RIGHT
                c = ws.cell(row=row, column=6, value=self._fmt(amt)); c.font = self.BOLD_FONT; c.border = self.BORDER
                if fill: c.fill = fill
                row += 1

            for c, w in [('A', 5), ('B', 35), ('C', 15), ('D', 15), ('E', 12), ('F', 18)]:
                ws.column_dimensions[c].width = w

            return self._save(wb)
        except Exception as e:
            import logging; logging.getLogger(__name__).error(f"Purchase Excel error: {e}")
            return None

    # ==================== DAILY REPORT EXCEL ====================

    def generate_daily_report(self, tenant_name: str, data: dict, report_date: date) -> bytes:
        try:
            wb = Workbook()
            f = lambda v: float(v or 0)

            # Sheet 1: Summary
            ws = wb.active
            ws.title = "Umumiy"
            ws['A1'] = f"{tenant_name} — KUNLIK HISOBOT"
            ws['A1'].font = Font(bold=True, size=16)
            ws['A2'] = f"Sana: {report_date.strftime('%d.%m.%Y')}"

            self._header_row(ws, 4, ["Ko'rsatkich", "Qiymat"])
            rows = [
                ("Sotuvlar soni", f"{data.get('total_sales_count', 0)} ta"),
                ("Jami summa", f"{f(data.get('total_amount', 0)):,.0f} so'm"),
                ("To'langan", f"{f(data.get('total_paid', 0)):,.0f} so'm"),
                ("Qarz", f"{f(data.get('total_debt', 0)):,.0f} so'm"),
                ("", ""),
                ("Naqd pul", f"{f(data.get('cash_amount', 0)):,.0f} so'm"),
                ("Plastik karta", f"{f(data.get('card_amount', 0)):,.0f} so'm"),
                ("O'tkazma", f"{f(data.get('transfer_amount', 0)):,.0f} so'm"),
                ("", ""),
                ("Umumiy qarzdorlik", f"{f(data.get('total_all_debt', 0)):,.0f} so'm"),
            ]
            for i, (k, v) in enumerate(rows, 5):
                ws.cell(row=i, column=1, value=k).border = self.BORDER
                ws.cell(row=i, column=2, value=v).border = self.BORDER

            ws.column_dimensions['A'].width = 25
            ws.column_dimensions['B'].width = 30

            # Sheet 2: Cashiers
            ws2 = wb.create_sheet("Kassirlar")
            self._header_row(ws2, 1, ["№", "Kassir", "Sotuvlar", "Jami", "To'langan", "Qarz"])
            for i, c in enumerate(data.get('cashiers', []), 1):
                ws2.cell(row=i+1, column=1, value=i).border = self.BORDER
                ws2.cell(row=i+1, column=2, value=c.get('name', '')).border = self.BORDER
                ws2.cell(row=i+1, column=3, value=c.get('sales_count', 0)).border = self.BORDER
                ws2.cell(row=i+1, column=4, value=f(c.get('total_amount', 0))).border = self.BORDER
                ws2.cell(row=i+1, column=5, value=f(c.get('paid_amount', 0))).border = self.BORDER
                ws2.cell(row=i+1, column=6, value=f(c.get('debt_amount', 0))).border = self.BORDER

            # Sheet 3: Low stock
            low = data.get('low_stock', [])
            if low:
                ws3 = wb.create_sheet("Kam qolgan")
                self._header_row(ws3, 1, ["№", "Tovar", "Qoldiq", "Birlik"])
                for i, s in enumerate(low, 1):
                    ws3.cell(row=i+1, column=1, value=i).border = self.BORDER
                    ws3.cell(row=i+1, column=2, value=s.get('name', '')).border = self.BORDER
                    ws3.cell(row=i+1, column=3, value=s.get('quantity', 0)).border = self.BORDER
                    ws3.cell(row=i+1, column=4, value=s.get('uom', '')).border = self.BORDER

            return self._save(wb)
        except Exception as e:
            import logging; logging.getLogger(__name__).error(f"Daily Excel error: {e}")
            return None
